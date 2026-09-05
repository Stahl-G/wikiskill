"""SpreadsheetBench adapter (verified_400 = paper's 80/40/280, bash+Python tools).

Data: ``dataset.json`` (400 items) plus per-task directories holding
``*_init.xlsx`` (agent input), ``*_golden.xlsx`` (target), and ``prompt.txt``.
Scoring follows the official cell-region contract: exact match on every cell of
``answer_position`` (e.g. ``A3:D32``) in ``answer_sheet`` between the agent's
edited workbook and the golden workbook. A missing agent sheet scores 0 on
that region (wrong workbook). A missing golden sheet is scorer infra and
raises. The agent is instructed to write concrete values (not formulas) into
the region so no spreadsheet engine is needed at scoring time; this deviation
from formula-preserving edits is recorded in the benchmark README.
"""

from __future__ import annotations

import json
import random
import re
from dataclasses import dataclass
from pathlib import Path

SYSTEM_PROMPT = """You are a spreadsheet expert who can manipulate spreadsheets through Python code.

{skill_section}

You need to solve the given spreadsheet manipulation question, which contains the following information:
- working_directory: The absolute path to your working directory where files are located. Make sure your output file is written inside the working_directory.
- instruction: The detailed description of the task.
- input_path: The spreadsheet file you need to manipulate, inside working_directory.
- output_path: The path of the output spreadsheet file, inside working_directory.

Do NOT create files outside the working_directory. Use the exact absolute paths provided. You have access to a bash tool that can execute any shell command and Python code. Write concrete values (not formulas) into the answer region of the output file."""


@dataclass(frozen=True)
class SpreadsheetCase:
    task_id: str
    instruction: str
    answer_position: str
    answer_sheet: str | None
    instruction_type: str
    init_name: str = ""
    golden_name: str = ""
    task_dir: str = ""

    @property
    def uid(self) -> str:
        return self.task_id


# Upstream malformations: these two items concatenate several sheet-qualified
# regions into one unparseable blob (no local sheet reference survives), so no
# region can be scored faithfully. Quarantined here, never silently dropped.
KNOWN_MALFORMED = frozenset({"41-47", "283-32"})


def load_cases(root: Path) -> list[SpreadsheetCase]:
    items = json.loads((root / "dataset.json").read_text())
    cases = []
    for item in items:
        if str(item["id"]) in KNOWN_MALFORMED:
            continue
        task_dir = root / item["spreadsheet_path"]
        files = {p.name for p in task_dir.iterdir() if p.suffix == ".xlsx"}
        golden = next((n for n in files if "golden" in n), "")
        init = next((n for n in files if "init" in n), "")
        cases.append(
            SpreadsheetCase(
                task_id=str(item["id"]),
                instruction=item["instruction"],
                answer_position=item["answer_position"],
                # Cell-Level tasks (275/400) carry no sheet fields; scoring
                # falls back to each workbook's first sheet.
                answer_sheet=item.get("answer_sheet"),
                instruction_type=item["instruction_type"],
                init_name=init,
                golden_name=golden,
                task_dir=str(task_dir),
            )
        )
    return cases


def split_cases(
    cases: list[SpreadsheetCase], seed: int = 20260904
) -> dict[str, list[str]]:
    """Deterministic split 80/40/rest over the 398 non-malformed tasks."""
    rng = random.Random(seed)
    pool = sorted(c.task_id for c in cases)
    rng.shuffle(pool)
    return {"train": pool[:80], "val": pool[80:120], "test": pool[120:]}


_CELL_RE = re.compile(r"([A-Z]+)(\d+)")
_RANGE_RE = re.compile(r"^([A-Z]+\d+):([A-Z]*\d+)$")
_COL_RANGE_RE = re.compile(r"^([A-Z]+):([A-Z]+)$")


def _col_to_idx(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _region(position: str, max_row: int | None = None) -> list[tuple[str, int]]:
    """Expand one reference to (column, row) cells.

    Accepts cell ranges (``A1:B2``), truncated ranges (``BD2:308`` keeps the
    start column), single cells (``B6``), and whole-column ranges (``A:G``,
    bounded by ``max_row``).
    """
    text = position.strip()
    col_range = _COL_RANGE_RE.match(text)
    if col_range is not None:
        if not max_row:
            raise ValueError(f"column range {text!r} needs max_row")
        c1i, c2i = _col_to_idx(col_range.group(1)), _col_to_idx(col_range.group(2))
        return [
            (col, row)
            for row in range(1, max_row + 1)
            for col in range(min(c1i, c2i), max(c1i, c2i) + 1)
        ]
    match = _RANGE_RE.match(text)
    if match is None:
        cell = _CELL_RE.fullmatch(text)
        if cell is None:
            raise ValueError(f"bad position {position!r}")
        col, row = _col_to_idx(cell.group(1)), int(cell.group(2))
        return [(col, row)]
    c1, r1 = _CELL_RE.fullmatch(match.group(1)).groups()
    r1 = int(r1)
    end = match.group(2)
    end_cell = _CELL_RE.fullmatch(end)
    if end_cell is None:  # truncated end like "308": keep start column
        c2, r2 = c1, int(end)
    else:
        c2, r2 = end_cell.group(1), int(end_cell.group(2))
    c1i, c2i = _col_to_idx(c1), _col_to_idx(c2)
    return [
        (col, row)
        for row in range(min(r1, r2), max(r1, r2) + 1)
        for col in range(min(c1i, c2i), max(c1i, c2i) + 1)
    ]


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_ref(text: str) -> bool:
    text = text.strip()
    return (
        _RANGE_RE.match(text) is not None
        or _COL_RANGE_RE.match(text) is not None
        or _CELL_RE.fullmatch(text) is not None
    )


def _split_top_level(position: str) -> list[str]:
    """Split on commas that sit outside single quotes."""
    segments, current, in_quote = [], [], False
    for ch in position:
        if ch == "'":
            in_quote = not in_quote
            current.append(ch)
        elif ch == "," and not in_quote:
            segments.append("".join(current))
            current = []
        else:
            current.append(ch)
    segments.append("".join(current))
    return [s for s in (seg.strip() for seg in segments) if s]


def _clean_sheet(name: str) -> str:
    return name.strip().strip("'").strip()


def parse_answer_regions(
    answer_sheet: str | None, answer_position: str
) -> list[tuple[str, str]]:
    """Tolerant parse of the dataset's answer fields.

    Raw values are inconsistent: sheet names may themselves contain commas
    (``'b2b, sez, de'``), ``answer_sheet`` may be comma-joined names, and
    ``answer_position`` may embed per-segment sheet prefixes with stray quotes
    (``Sheet1'!A1:F14``, ``'99250!A1:F9'``), single cells (``B6``), column
    ranges (``A:G``), or truncated ranges (``BD2:308``). Segments split only
    on commas outside quotes; each keeps its own sheet prefix when present,
    else falls back to the first parsed answer sheet. ~19/400 cases list
    extra sheets without positions; those secondary sheets are not scored
    (documented deviation).
    """
    default_sheet = None
    if answer_sheet:
        names = [_clean_sheet(x) for x in answer_sheet.split(",")]
        default_sheet = next((n for n in names if n), None)

    def parse_segment(segment: str) -> tuple[str, str] | None:
        segment = segment.strip()
        if not segment:
            return None
        if "!" in segment:
            sheet_part, _, ref_part = segment.rpartition("!")
            ref_part = ref_part.strip().strip("'")
            if _is_ref(ref_part):
                return (_clean_sheet(sheet_part) or default_sheet, ref_part)
        if _is_ref(segment):
            return (default_sheet, segment)
        return None

    regions = []
    for segment in _split_top_level(answer_position):
        parsed = parse_segment(segment)
        if parsed is not None:
            regions.append(parsed)
    if not regions:
        raise ValueError(f"bad position {answer_position!r}")
    return regions


def score_workbook(
    agent_output_path: Path,
    golden_path: Path,
    answer_sheet: str | None,
    answer_position: str,
) -> tuple[float, int, int]:
    """Cell-exact match over all answer regions; returns (score, matched, total).

    Both workbooks load with ``data_only=True``: goldens ship cached computed
    values, and the agent is instructed to write concrete values, so formula
    objects never reach the comparison.
    """
    from openpyxl import load_workbook

    out_wb = load_workbook(agent_output_path, data_only=True)
    gold_wb = load_workbook(golden_path, data_only=True)

    def sheet_or_none(wb, name: str | None):
        if name is None:
            return wb.worksheets[0]
        if name in wb.sheetnames:
            return wb[name]
        return None

    matched = 0
    total = 0
    for sheet, position in parse_answer_regions(answer_sheet, answer_position):
        gold_ws = sheet_or_none(gold_wb, sheet)
        if gold_ws is None:
            raise KeyError(f"golden missing sheet {sheet!r}")
        out_ws = sheet_or_none(out_wb, sheet)
        max_row = gold_ws.max_row if _COL_RANGE_RE.match(position) else None
        cells = _region(position, max_row)
        if out_ws is None:
            # Agent omitted the answer sheet: that is a wrong workbook, not
            # a scorer crash. Count the gold region as unmatched.
            total += len(cells)
            continue
        for col, row in cells:
            letter = ""
            c = col
            while c:
                c, rem = divmod(c - 1, 26)
                letter = chr(65 + rem) + letter
            total += 1
            if _norm(out_ws[f"{letter}{row}"].value) == _norm(
                gold_ws[f"{letter}{row}"].value
            ):
                matched += 1
    return (1.0 if total and matched == total else 0.0, matched, total)
